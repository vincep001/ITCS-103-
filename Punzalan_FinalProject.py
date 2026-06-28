import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op
import time

MENU ={
"Nilaga": 130,
"Kaldereta": 160,
"Sisig": 120,
"Adobo": 120,
"Sinigang": 140,
"Lumpia": 90,

"Coke": 20,
"Sprite": 20,
"Lemonade": 30,
"Nestea": 30,
"Royal": 20,
"Melon Juice": 30,

"Plain White Rice": 15,
"Java Rice": 15
}

cart = []
grand_total = 0
order_id = 1


#Kusina ni vince
#gawa ng excel file
#customer name, product, price, quanity


window = tk.Tk()
window.title("Kusina ni Vince")
window.geometry("1200x670")
window.resizable(False, False)
window.configure(bg="#D89C00")



logo = tk.PhotoImage(file="KusinaNivinceLogo.png")
logo_label = tk.Label(window, image=logo, bg="#D89C00")
logo_label.image = logo  
logo_label.place(x=10, y=5)


time_label = tk.Label(window, font=("Poppins", 14, "bold"), bg="#D89C00", fg="white")
time_label.place(x=1070, y=10)

def update_time():
    current_time = time.strftime("%I:%M:%S %p")
    time_label.config(text=current_time)
    time_label.after(1000, update_time) 

window.grid_rowconfigure(1, weight=1)
window.grid_rowconfigure(2, weight=1)
window.grid_columnconfigure(0, weight=1)
window.grid_columnconfigure(1, weight=1)
window.grid_rowconfigure(3, weight=1)




try:
    workbook = op.load_workbook("restoproj.xlsx")

except FileNotFoundError:

    workbook = op.Workbook()
    sheet = workbook.active

    sheet['A1'] = "Order ID"
    sheet['B1'] = "Customer Name"
    sheet['C1'] = "Product"
    sheet['D1'] = "Quantity"
    sheet['E1'] = "Price"
    sheet['F1'] = "Total"

    workbook.save("restoproj.xlsx")




#FUNCTIONSSSSSSSS

def display_excel():

    workbook = op.load_workbook("restoproj.xlsx")
    sheet = workbook.active

    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)





def input_validation():

    cust = cname_entry.get()
    prod = order_entry.get()
    quan = qty_entry.get()
    pric = price_entry.get()
   
    if not cust or not prod or not quan or not pric:
        messagebox.showerror("Error", "Please fill up all the fields!")
        return False
    
    if not quan.isdigit() or not pric.isdigit():
        messagebox.showerror("Error", "Quantity and Price must be numbers:(")
        return False
    
    if prod.title() not in MENU:
        messagebox.showerror("Error", "Product is not in the menu!")
        return False
    return True



def add_item():

    global grand_total

    if not input_validation():
        return
    
    cust = cname_entry.get()
    prod = order_entry.get().strip().title()
    quan = int(qty_entry.get())
    pric = int(price_entry.get())

    total = quan * pric

    cart.append((cust, prod, quan, pric, total))

    receipt_text.insert(tk.END, f"{prod} x{quan} = ₱{total}\n")

    grand_total += total
    total_label.config(text=f"TOTAL: ₱{grand_total}")

    order_entry.delete(0, tk.END)
    qty_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)


def submit_order():
    global order_id

    if not cart:
        messagebox.showerror("Error", "No items in order!")
        return

    workbook = op.load_workbook("restoproj.xlsx")
    sheet = workbook.active

    current_id = order_id

    for item in cart:
        cust, prod, qty, price, total = item
        sheet.append([current_id, cust, prod, qty, price, total])


    workbook.save("restoproj.xlsx")

    messagebox.showinfo("Success", f"Order {current_id} saved successfully!")

    order_id += 1


    cart.clear()
    receipt_text.delete("1.0", tk.END)

    global grand_total
    grand_total = 0
    total_label.config(text="TOTAL: ₱0")

    display_excel()







def auto_populate(event):

    selected = table.focus()

    values = table.item(selected, "values")
    if values:

        cname_entry.delete(0, tk.END)
        order_entry.delete(0, tk.END)
        qty_entry.delete(0, tk.END)
        price_entry.delete(0, tk.END)

        cname_entry.insert(0, values[1])
        order_entry.insert(0, values[2])
        qty_entry.insert(0, values[3])
        price_entry.insert(0, values[4])

#deletes
def delete_data():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return
    
    values = table.item(selected, "values")

    record_id = str(values[0])

    confirm = messagebox.askyesno("Confirm", "Delete this order?")

    if not confirm:
        return
    
    workbook = op.load_workbook("restoproj.xlsx")
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == record_id:
            sheet.delete_rows(i)
            break

    workbook.save("restoproj.xlsx")
    messagebox.showinfo("Success", "Order deleted successfully")

    display_excel()



def update_data():

    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return

    if not input_validation():
        return

    values = table.item(selected, "values")

    record_id = int(values[0])

    cust = cname_entry.get()
    prod = order_entry.get()
    quan = int(qty_entry.get())
    pric = int(price_entry.get())

    total = quan * pric

    workbook = op.load_workbook("restoproj.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):

        if str(row[0].value) == str(record_id):

            row[1].value = cust
            row[2].value = prod
            row[3].value = quan
            row[4].value = pric
            row[5].value = total
            break
    workbook.save("restoproj.xlsx")

    messagebox.showinfo("Success", "Order updated successfully!")

    display_excel()



def open_menu():

    menu_window = tk.Toplevel(window)
    menu_window.title("Kusina ni vince Menu")
    menu_window.geometry("400x500")
    menu_window.resizable(False, False)
    menu_window.configure(bg="#D89C00")

    #bg
    bg = tk.PhotoImage(file="BG_ULAM.png")

    bg_label = tk.Label(menu_window, image=bg)
    bg_label.image = bg
    bg_label.place(x=0, y=0)

    # MGA ULAM
    mga_ulam = tk.Label(menu_window, text="MGA ULAM:", font=("Poppins", 16, "bold"), bg="#D89C00")
    mga_ulam.place(x=30, y=20)

    nilaga = tk.Label(menu_window, text="Nilaga ---------------- ₱130", font=("Poppins", 12, "bold"), bg="#D89C00")
    nilaga.place(x=110, y=60)

    kaldereta = tk.Label(menu_window, text="Kaldereta ----------- ₱160", font=("Poppins", 12, "bold"), bg="#D89C00")
    kaldereta.place(x=110, y=85)

    sisig = tk.Label(menu_window, text="Sisig ------------------ ₱120", font=("Poppins", 12, "bold"), bg="#D89C00")
    sisig.place(x=110, y=110)

    adobo = tk.Label(menu_window, text="Adobo --------------- ₱120", font=("Poppins", 12, "bold"), bg="#D89C00")
    adobo.place(x=110, y=135)

    sinigang = tk.Label(menu_window, text="Sinigang ------------- ₱140", font=("Poppins", 12, "bold"), bg="#D89C00")
    sinigang.place(x=110, y=160)

    lumpia = tk.Label(menu_window, text="Lumpia ---------------- ₱90", font=("Poppins", 12, "bold"), bg="#D89C00")
    lumpia.place(x=110, y=185)

    #DRINKS
    mga_inumin = tk.Label(menu_window, text="MGA INUMIN:", font=("Poppins", 16, "bold"), bg="#D89C00")
    mga_inumin.place(x=30, y=215)

    coke = tk.Label(menu_window, text="Coke ------------------ ₱20", font=("Poppins", 12, "bold"), bg="#D89C00")
    coke.place(x=110, y=250)

    sprite = tk.Label(menu_window, text="Sprite ---------------- ₱20", font=("Poppins", 12, "bold"), bg="#D89C00")
    sprite.place(x=110, y=275)

    lemonade = tk.Label(menu_window, text="Lemonade ------------ ₱30", font=("Poppins", 12, "bold"), bg="#D89C00")
    lemonade.place(x=110, y=300)

    nestea = tk.Label(menu_window, text="Nestea ---------------- ₱30", font=("Poppins", 12, "bold"), bg="#D89C00")
    nestea.place(x=110, y=325)

    royal = tk.Label(menu_window, text="Royal ----------------- ₱20", font=("Poppins", 12, "bold"), bg="#D89C00")
    royal.place(x=110, y=350)

    melon = tk.Label(menu_window, text="Melon Juice --------- ₱30", font=("Poppins", 12, "bold"), bg="#D89C00")
    melon.place(x=110, y=375)

    #mga kanin
    mga_kanin = tk.Label(menu_window, text="MGA KANIN:", font=("Poppins", 16, "bold"), bg="#D89C00")
    mga_kanin.place(x=30, y=405)

    plain_rice = tk.Label(menu_window, text="Plain White Rice ----- ₱15", font=("Poppins", 12, "bold"), bg="#D89C00")
    plain_rice.place(x=110, y=440)

    java_rice = tk.Label(menu_window, text="Java Rice ------------- ₱15", font=("Poppins", 12, "bold"), bg="#D89C00")
    java_rice.place(x=110, y=465)

   






#tkinter widgets
title = tk.Label(window, text="Kusina ni Vince", font=("Segoe UI", 24, "bold"), bg="#D89c00", fg="white")
title.grid(row=0, column=0, columnspan=2, pady=10)


lframe = tk.Frame(window, bg="#D89c00", width=450, height=250)
lframe.grid(row=1, column=0, padx=(30, 10), pady=40, sticky="nw")


rframe = tk.Frame(window, bg="#D89c00", width=450, height=250)
rframe.grid(row=1, column=1, padx=(10, 30), pady=10, sticky="nw")





#customer
cname_entry = tk.Entry(lframe, font=("Poppins", 14), width=25)
cname_entry.grid(row=0, column=1, pady=8, padx=15)

cname_label = tk.Label(lframe, text="Customer Name:", font=("Poppins", 14, "bold"), bg="#D89c00", fg="black", width=18, anchor="e")
cname_label.grid(row=0, column=0, padx=5, pady=5)


#order
order_entry = tk.Entry(lframe, font=("Poppins", 14), width=25)
order_entry.grid(row=1, column=1, pady=8, padx=15)

order_label = tk.Label(lframe, text="Food/Drinks/Rice:", font=("Poppins", 14, "bold"), bg="#d89c00", fg="black", width=18, anchor="e")
order_label.grid(row=1, column=0, padx=5, pady=5)


#price
price_entry = tk.Entry(lframe, font=("poppins", 14), width=25)
price_entry.grid(row=2, column=1, pady=8, padx=15)


price_label = tk.Label(lframe, text="Price:", font=("poppins", 14, "bold"),bg="#d89c00", fg="black", width=18, anchor="e")
price_label.grid(row=2, column=0, padx=5, pady=5)

#quanti
qty_entry = tk.Entry(lframe, font=("poppins", 14), width=25)
qty_entry.grid(row=3, column=1, pady=8, padx=15)


qty_label = tk.Label(lframe, text="Quantity:", font=("poppins", 14, "bold"), bg="#d89c00", fg="black", width=18, anchor="e")
qty_label.grid(row=3, column=0, padx=5, pady=5)



#receptis
receipt_label = tk.Label(rframe, text="RECEIPT", font=("Poppins", 16, "bold"), bg="#D89c00")
receipt_label.grid(row=0, column=0, columnspan=2)

receipt_text = tk.Text(rframe, height=10, width=30)
receipt_text.grid(row=1, column=0, columnspan=2)

total_label = tk.Label(rframe, text="TOTAL: ₱0", bg="#D89c00", font=("Poppins", 14, "bold"))
total_label.grid(row=2, column=0, columnspan=2)


#buttons

button_frame = tk.Frame(window, bg="#D89c00")
button_frame.grid(row=2, column=0, columnspan=2, pady=10)

add_btn = tk.Button(button_frame, text="ADD ITEM", font=("Poppins", 12, "bold"), bg="gray", width=12, command=add_item)
add_btn.grid(row=0, column=0, padx=5)

update_btn = tk.Button(button_frame, text="Update", font=("Poppins", 12, "bold"), bg="gray", command=update_data, width=12)
update_btn.grid(row=0, column=1, padx=5)

delete_btn = tk.Button(button_frame, text="Delete", bg="red", fg="white", font=("Poppins", 12, "bold"), command=delete_data, width=12)
delete_btn.grid(row=0, column=2, padx=5)

menu_btn = tk.Button(button_frame, text="Menu", font=("Poppins", 12, "bold"), bg="lightblue", command=open_menu, width=12)
menu_btn.grid(row=0, column=3, padx=5)

submitorder_btn = tk.Button(button_frame, text="SUBMIT ORDER", font=("Poppins", 12, "bold"), bg="green", fg="white", command=submit_order, width=12)
submitorder_btn.grid(row=0, column=4,padx=5)


table = ttk.Treeview(window, columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"), show="headings", height=13)

for col in("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"):
    table.heading(col, text=col, anchor="center")

    table.column("Order ID", width=80, anchor="center")
    table.column("Customer Name", width=200, anchor="center")
    table.column("Product", width=200, anchor="center")
    table.column("Quantity", width=100, anchor="center")
    table.column("Price", width=100, anchor="center")
    table.column("Total", width=120, anchor="center")

table.grid(row=3, column=0, columnspan=2, padx=20, pady=(10,20), sticky="nsew")


#add auto_populate function mamaya
table.bind("<<TreeviewSelect>>", auto_populate)

#display_excel function mamaya
display_excel()














update_time()  

window.mainloop()