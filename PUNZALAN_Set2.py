import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op


window = tk.Tk()
window.title("Simple Ordering System")
window.configure(bg="lightblue")


#workbook

# workbook = op.Workbook()
# sheet = workbook.active

# sheet['A1'] = "Order ID"
# sheet['B1'] = "Customer Name"
# sheet['C1'] = "Product"
# sheet['D1'] = "Quantity"
# sheet['E1'] = "Price"
# sheet['F1'] = "Total"

# workbook.save("ordersDB.xlsx")

#functions
#display


def display_excel():
    workbook = op.load_workbook("ordersDB.xlsx")
    sheet = workbook.active

    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)




#validate
def input_validation():
    cust = cname_entry.get()
    prod = product_entry.get()
    quan = qty_entry.get()
    pric = price_entry.get()

    if not cust or not prod or not quan or not pric:
        messagebox.showerror("Error", "Required fill to the following input")
        return False
    
    if not pric.isdigit() or not quan.isdigit():
        messagebox.showerror("Error", "Need to be number!")
        return False
    
    return True

#saving, append
def saving():
    if not input_validation():
        return
    cust = cname_entry.get()
    prod = product_entry.get()
    quan = int(qty_entry.get())
    pric = int(price_entry.get())
    total = quan * pric

    workbook = op.load_workbook("ordersDB.xlsx")
    sheet = workbook.active

    new = sheet.max_row

    sheet.append([new,cust, prod, quan, pric, total])
    workbook.save("ordersDB.xlsx")

    cname_entry.delete(0, tk.END)
    product_entry.delete(0, tk.END)
    qty_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)

    messagebox.showinfo("Success", "Record info has been ")
    display_excel()



#auto_populate
def auto_populate(event):
    selected = table.focus()
    values = table.item(selected, "values")

    if values:
        cname_entry.delete(0, tk.END)
        product_entry.delete(0, tk.END)
        qty_entry.delete(0, tk.END)
        price_entry.delete(0, tk.END)

        cname_entry.insert(0, values[1])
        product_entry.insert(0, values[2])
        qty_entry.insert(0, values[3])
        price_entry.insert(0, values[4])

def delete_data():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return
    
    values = table.item(selected, "values")
    record_id = str(values[0])  

    confirm = messagebox.askyesno("Confirm", "Are you sure you want to delete this record?")
    if not confirm:
        return
    
    workbook = op.load_workbook("ordersDB.xlsx")
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == record_id:
            sheet.delete_rows(i)
            break
    
    workbook.save('ordersDB.xlsx')

    messagebox.showinfo("Success", "Record deleted successfully!")
    
    display_excel()



#update
def update_data():
       selected = table.focus()

       if not selected:
           messagebox.showerror("Error", "Select a record first!")
           return
       if not input_validation():
           return
       
       values = table.item(selected, "values")
       record_id = (values[0])

       cust = cname_entry.get()
       prod = product_entry.get()
       quan = int(qty_entry.get())
       pric = int(price_entry.get())
       total = quan * pric

       workbook = op.load_workbook("ordersDB.xlsx")
       sheet = workbook.active

       for row in sheet.iter_rows(min_row=2):
           if row[0].value == record_id:
               row[1].value = cust
               row[2].value = prod
               row[3].value = quan
               row[4].value = pric
               row[5].value = total
       workbook.save('ordersBD.xlsx')

       messagebox.showinfo("Success", "Record updated successfully!")

       


    
       







# Form Title
title = tk.Label(window, text="Simple Ordering System", font=("Times New Roman", 14, "bold"), bg="lightblue")
title.grid(row=0, column=0, columnspan=6)

# Frame
genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)

# Customer Name Entry
cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = tk.Label(genframe, text="Customer Name", font=("Poppins", 10, "italic"), bg="lightblue")
cname_label.grid(row=3, column=1, columnspan=2)

# Product Entry
product_entry = tk.Entry(genframe, font=("Poppins", 12))
product_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

product_label = tk.Label(genframe, text="Product", font=("Poppins", 10, "italic"), bg="lightblue")
product_label.grid(row=3, column=3, columnspan=2)

# Quantity Entry
qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

qty_label = tk.Label(genframe, text="Quantity", font=("Poppins", 10, "italic"), bg="lightblue")
qty_label.grid(row=5, column=1, columnspan=2)

# Price Entry
price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

price_label = tk.Label(genframe, text="Price", font=("Poppins", 10, "italic"), bg="lightblue")
price_label.grid(row=5, column=3, columnspan=2)

# Buttons
submit_btn = tk.Button(window, text="Submit", font=("Poppins", 12, "bold"), bg="lightpink", command=saving)
submit_btn.grid(row=6, column=1, pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Poppins", 12, "bold"), bg="lightgreen", command=update_data)
update_btn.grid(row=6, column=2)

delete_btn = tk.Button(window, text="Delete", bg="red", fg="white",font=("Poppins", 12, "bold"), command=delete_data)
delete_btn.grid(row=6, column=3)

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"),
    show="headings"
)

for headings in ("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", auto_populate)

display_excel()

window.mainloop()